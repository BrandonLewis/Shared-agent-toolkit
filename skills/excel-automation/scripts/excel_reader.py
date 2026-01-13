#!/usr/bin/env python3
"""
Excel Reader Utility
Reads data from Excel files and outputs to JSON or CSV
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is not installed. Install with: pip install openpyxl")
    sys.exit(1)


def parse_range(range_str):
    """Parse Excel range string like 'A1:D10' into start/end cells"""
    if ':' in range_str:
        start, end = range_str.split(':')
        return start, end
    return range_str, range_str


def read_excel(file_path, sheet_name=None, cell_range=None, has_header=True):
    """
    Read data from Excel file

    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read (default: active sheet)
        cell_range: Cell range to read (e.g., 'A1:D10')
        has_header: Whether first row contains headers

    Returns:
        Dictionary with data and metadata
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Determine range to read
        if cell_range:
            start_cell, end_cell = parse_range(cell_range)
            cells = ws[cell_range]
        else:
            cells = ws.iter_rows(values_only=True)

        # Extract data
        data = []
        headers = None

        for i, row in enumerate(cells):
            row_values = [cell for cell in row]

            if i == 0 and has_header:
                headers = row_values
            else:
                if headers:
                    data.append(dict(zip(headers, row_values)))
                else:
                    data.append(row_values)

        return {
            'success': True,
            'file': str(file_path),
            'sheet': ws.title,
            'rows': len(data),
            'columns': len(headers) if headers else len(data[0]) if data else 0,
            'headers': headers,
            'data': data
        }

    except FileNotFoundError:
        return {'success': False, 'error': f'File not found: {file_path}'}
    except KeyError:
        return {'success': False, 'error': f'Sheet not found: {sheet_name}'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def main():
    parser = argparse.ArgumentParser(description='Read data from Excel files')
    parser.add_argument('--file', required=True, help='Path to Excel file')
    parser.add_argument('--sheet', help='Sheet name (default: active sheet)')
    parser.add_argument('--range', help='Cell range to read (e.g., A1:D10)')
    parser.add_argument('--no-header', action='store_true', help='First row is not a header')
    parser.add_argument('--output', choices=['json', 'csv'], default='json', help='Output format')

    args = parser.parse_args()

    result = read_excel(
        file_path=args.file,
        sheet_name=args.sheet,
        cell_range=args.range,
        has_header=not args.no_header
    )

    if not result['success']:
        print(f"Error: {result['error']}", file=sys.stderr)
        sys.exit(1)

    if args.output == 'json':
        print(json.dumps(result, indent=2, default=str))
    else:
        # CSV output
        if result['headers']:
            print(','.join(str(h) for h in result['headers']))
        for row in result['data']:
            if isinstance(row, dict):
                print(','.join(str(v) for v in row.values()))
            else:
                print(','.join(str(v) for v in row))


if __name__ == '__main__':
    main()
